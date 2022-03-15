'''
test xlrd , openpyxl , pandas read XLSX file

result : 
fail : xlrd
success : openpyxl , pandas

1. test xlrd
2. test openpyxl 
3. test pandas 

4. example of reading / writing xlsx file with pandas
'''
# fail
def Test_xlrd(): 

    #Import the xlrd module
    import xlrd

    # Open the Workbook
    workbook = xlrd.open_workbook("DATA/test0.xlsx")

    # Open the worksheet
    worksheet = workbook.sheet_by_index(0)

    # Iterate the rows and columns
    for i in range(0, 2):
        for j in range(0, 2):
            # Print the cell values with tab space
            print(worksheet.cell_value(i, j), end='\t')
        print('')

# success
def Test_openpyxl():

    import openpyxl
    wookbook = openpyxl.load_workbook("DATA/test0.xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            print(col[i].value, end="\t\t")
        print('')

# success
def Test_pandas():
    # Import pandas
    import pandas as pd

    # Load the xlsx file
    excel_data = pd.read_excel('DATA/test0.xlsx')
    # Read the values of the file in the dataframe
    # Print the content
    print("The content of the file is:\n", excel_data)

def Demo_Pandas():
    import pandas as pd
    # read xlsx file
    excel_data = pd.read_excel('DATA/test0.xlsx')
    # write xlsx file
    excel_data.to_excel('DATA/test1.xlsx')